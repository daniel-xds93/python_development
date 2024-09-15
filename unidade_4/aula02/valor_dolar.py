from selenium import webdriver

from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By

# a linha abaixo importa biblioteca para trabalhar com aplanilha
from openpyxl import load_workbook

navegador = webdriver.Chrome()

# a linha abaixo acessa um site
navegador.get('https://www.google.com')

# a linha abaixo seleciona um elemento na pagia e cola o texto indicado
navegador.find_element(By.XPATH, '//*[@id="APjFqb"]').send_keys('cotação dolar')

navegador.find_element(By.XPATH, '/html/body/div[1]/div[3]/form/div[1]/div[1]/div[4]/center/input[1]').send_keys(Keys.ENTER)

valor_dolar = navegador.find_element(By.XPATH, '//*[@id="knowledge-currency__updatable-data-column"]/div[3]/div/div[2]/input').get_attribute('value')

print(f'Valor atual do dolar R$ {valor_dolar}')

# a linha abaixo abre o arquivo
arquivo = load_workbook('produtos.xlsx')

# a linha abaixo retorna os nomes das planilhas no arquivo
#print(arquivo.sheetnames)

plan = arquivo['lista']

valor_dolar = valor_dolar.replace('.', ',')

plan.cell(row=1, column=2).value = valor_dolar

arquivo.save('produtos.xlsx')

print("planilha atualizada com sucesso!!!")